from django.shortcuts import render, redirect,get_object_or_404
from .models import *
from .forms import *
from django.contrib import messages
from django.http import JsonResponse
from django.contrib.sessions.models import Session
from django.db import connection
from django.core import serializers
from django.http import HttpResponse
from django.db.models import Count
from django.db.models import Sum
from datetime import datetime
from django.http import HttpResponseRedirect
from django.conf.urls import url
import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import itertools    


def pagination_cnt(request):
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    sql = '''SELECT school_schoolsetting.pagination_count from school_schoolsetting WHERE 
    school_schoolsetting.school_id_id='%d' ''' % (school_id)
    post = cursor.execute(sql)
    row = cursor.fetchone()
    count = row[0]
    return count
def export_subject_to_xlsx(request):
    if request.method == 'GET':
        class_id = request.GET.get('cls_id')
        section_id = request.GET.get('sec_id')
        count = request.GET.get('count_val')
        staff_id = request.session['user_id']
        a = int(class_id)
        b = int(section_id)
        cnt = int(count)
        subject_queryset = AssignSubjectTeacher.objects.filter(class_id=a,section_id=b,staff_id=staff_id).order_by('subject_id')
        cursor = connection.cursor()
        post = '''SELECT  school_studentdetail.register_number,school_studentdetail.student_name from school_studentsection INNER JOIN school_studentdetail
        ON school_studentsection.student_id_id=school_studentdetail.id where 
        school_studentsection.class_id_id='%d' AND school_studentsection.section_id_id='%d' order by school_studentsection.student_id_id ''' % (a,b)
        query = cursor.execute(post)
        row = cursor.fetchall()
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-marksheet.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Mark Sheet'

        # Define the titles for columns
        columns = [
            'Student Register Number',
           
        ]

        name = ["0", "Student Name"]

        num = ["0","S.No"]

        for a in range(1,cnt+1):
            num.append(a)

        for i in row:
            columns.append(i[0])
            name.append(i[1])
            
        for i in range(1,cnt+2):
            worksheet.cell(row=i, column=1).value = num[i]

        for i in range(1,cnt+2):
            worksheet.cell(row=i, column=2).value = name[i]

        row_num = 3

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=col_num, column=row_num)
            cell.value = column_title

        # Iterate through all movies
        for sub in subject_queryset:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                sub.subject_id.subject_name,
                
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=col_num, column=row_num)
                cell.value = cell_value 

        workbook.save(response)

        return response      
def adminlogin(request):
    if request.session.has_key('username'):
        return redirect("dashboard")
    else:
        if request.method == 'POST':
            username = request.POST['username']
            password =  request.POST['password']
            post = SchoolDetail.objects.filter(username=username,password=password)
            if post:
                username = request.POST['username']
                request.session['username'] = username
                a = request.session['username']
                sess = SchoolDetail.objects.only('id').get(username=a).id
                request.session['schoolname']=sess
                return redirect("dashboard")
            else:
                messages.success(request, 'Invalid Username or Password')
    return render(request, 'index.html', {})
def teacher_login(request):
    if request.session.has_key('teacher'):
        return redirect("teacher_dashboard")
    else:
        if request.method == 'POST':
            username = request.POST['username']
            password =  request.POST['password']
            post = StaffDetail.objects.filter(username=username,password=password)
            if post:
                username = request.POST['username']
                request.session['teacher'] = username
                a = request.session['teacher']
                user_id = StaffDetail.objects.only('staff_id').get(username=a).staff_id
                request.session['user_id']=user_id
                sess = StaffDetail.objects.only('school_id_id').get(staff_id=user_id).school_id_id
                request.session['schoolname']=sess
                return redirect("teacher_dashboard")
            else:
                messages.success(request, 'Invalid Username or Password')
    return render(request, 'teacher_login.html', {})
def student_login(request):
    if request.session.has_key('student'):
         return redirect("student_dashboard")
    else:
        if request.method == 'POST':
            username = request.POST['username']
            password =  request.POST['password']
            post = StudentDetail.objects.filter(username=username,password=password)
            if post:
                username = request.POST['username']
                request.session['student'] = username
                a=request.session['student']
                stud_id = StudentDetail.objects.only('id').get(username=a).id
                request.session['student_id']=stud_id
                stud_reg_no = StudentDetail.objects.only('register_number').get(username=a).register_number
                request.session['register_number']=stud_reg_no
                sess = StudentDetail.objects.only('school_id_id').get(username=a).school_id_id
                request.session['schoolname']=sess
                return redirect("student_dashboard")
            else:
                messages.success(request, 'Invalid Username or Password')
    return render(request, 'student_login.html', {})
def dashboard(request):
	if request.session.has_key('username'):
		posts = request.session['username']
		query = SchoolDetail.objects.filter(username=posts)
		return render(request, 'dashboard.html', {"query":query})
	else:
		return render(request, 'index.html',{})
def teacher_dashboard(request):
	return render(request, 'teacher_dashboard.html',{})
def student_dashboard(request):
	return render(request, 'student_dashboard.html',{})
def logout(request):
    try:
        Session.objects.all().delete()
    except:
     pass
    return render(request, 'index.html', {})

def teacher_logout(request):
    try:
        Session.objects.all().delete()
    except:
     pass
    return render(request, 'teacher_login.html', {})

def student_logout(request):
    try:
        Session.objects.all().delete()
    except:
     pass
    return render(request, 'student_login.html', {})

def menu(request):
    if request.session.has_key('username'):
        posts = request.session['username']
        query = SchoolDetail.objects.filter(username=posts)
        data = {
            "user":user
        }
        return JsonResponse(data)
def add_teacher(request):
    if request.method == "POST":
        form = StaffForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect('teachers')
    else:
        form = StaffForm()
    if "GET" == request.method:
        return render(request, 'add_teacher.html', {'forms':form})
    else:
        file = request.FILES['excel_file']
        school_id = request.session['schoolname']
        ids = SchoolDetail.objects.get(id=int(school_id))
        workbook = openpyxl.load_workbook(file, read_only=True)
        # Get name of the first sheet and then open sheet by name
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)

        data = []
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
            user_name = row[5].value
            user_exist = StaffDetail.objects.filter(username=user_name)
            if user_exist:
                messages.success(request, user_name + " Username Already Exist")
            else:
                staff = StaffDetail()
                staff.staff_name = row[0].value
                staff.mobile = row[1].value
                staff.degree = row[2].value
                staff.designation = row[3].value
                staff.age = row[4].value
                staff.username = row[5].value
                staff.password = row[6].value
                staff.school_id = ids
                data.append(staff)
        # Bulk create data
        StaffDetail.objects.bulk_create(data)
        return redirect("teachers")
    return render(request, 'add_teacher.html', {'forms':form})
def teachers(request):
    school_id = request.session['schoolname']
    teacher = StaffDetail.objects.filter(school_id_id=school_id)
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(teacher,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    return render(request, 'teachers.html', {'teacher':users})
def add_student(request):
    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect('students')
    else:
        form = StudentForm()
    if "GET" == request.method:
        return render(request, 'add_students.html', {'form':form})
    else:
        file = request.FILES['excel_file']
        school_id = request.POST.get('school_id')
        ids = SchoolDetail.objects.get(id=int(school_id))
        workbook = openpyxl.load_workbook(file, read_only=True)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)

        data = []
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
            user_name = row[5].value
            user_exist = StudentDetail.objects.filter(username=user_name)
            if user_exist:
                messages.success(request, user_name + " Username Already Exist")
            else:
                student = StudentDetail()
                student.student_name = row[0].value
                student.register_number = row[1].value
                student.admission_number = row[2].value
                student.emergency_number = row[3].value
                student.date_of_birth = row[4].value
                student.username = row[5].value
                student.password = row[6].value
                student.school_id = ids
                data.append(student)
        StudentDetail.objects.bulk_create(data)
        return redirect("students")
    return render(request, 'add_students.html', {'form':form})
def edit_teacher(request,pk):
    post = get_object_or_404(StaffDetail, pk=pk)
    if request.method == "POST":
        form = StaffEditForm(request.POST,instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect('teachers')
    else:
        form = StaffEditForm(instance=post)
    return render(request, 'edit_teacher.html', {'form': form})
def teacher_delete(request, pk):
    student = StaffDetail.objects.get(staff_id=pk)
    student.delete()
    return redirect("teachers")
def students(request):
    ids = StudentDetail.objects.latest('id')
    school_id = request.session['schoolname']
    student = StudentDetail.objects.filter(school_id_id=school_id).order_by('-id')
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(student,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    return render(request, 'students.html', {'student':users,'ids':ids})
def student_edit(request,pk):
    post = get_object_or_404(StudentDetail, pk=pk)
    if request.method == "POST":
        form = StudentEditForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect("students")
    else:
        form = StudentEditForm(instance=post)
    return render(request, 'student_edit.html', {'form': form})
def student_delete(request, pk):
    student = StudentDetail.objects.get(id=pk)
    student.delete()
    return redirect("students")
def add_class_section(request):
    class_id = dict()
    school_id = request.session['schoolname']
    academic_year = request.session['academic_year']
    cls_name = Section.objects.filter(class_id__in=Class.objects.filter(academic_year=academic_year,school_id=school_id))
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(cls_name,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    if request.method == "POST":
        text = request.POST.get('class_name')
        class_id = text[text.index("-")]

    return render(request, 'add_classes_sections.html', {'cls_name':users,'class_id':class_id})
def edit_classes_sections(request,cls_id,sec_id):
    if request.method == "POST":
        cls_name = request.POST.get('class_name')
        sec_name = request.POST.get('section_name')
        class_update = Class.objects.filter(class_id=cls_id).update(class_name=cls_name)
        section_update = Section.objects.filter(section_id=sec_id).update(section_name=sec_name)
        return redirect("add_class_section")
    post = cls_name = Section.objects.filter(section_id=sec_id,class_id__in=Class.objects.filter(class_id=cls_id))
    return render(request, 'edit_classes_sections.html', {'post':post})
def delete_classes_sections(request,cls_id,sec_id):
    class_del = Class.objects.get(class_id=cls_id)
    class_del.delete()
    section_del = Section.objects.get(section_id=sec_id)
    section_del.delete()
    return redirect("add_class_section")
def add_subject(request):
    school_id = request.session['schoolname']
    academic_year = request.session['academic_year']
    row = Subject.objects.filter(school_id=school_id).select_related('class_id').select_related('section_id').order_by('subject_id')
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(row,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    if request.method == "POST":
        form = SubjectForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            messages.success(request, 'Subject Added Successfully')
            form = SubjectForm()
    else:
        form = SubjectForm()

    return render(request, 'add_subject.html', {'row':users,'form':form})
def edit_subject(request,pk):
    post = get_object_or_404(Subject, pk=pk)
    if request.method == "POST":
        form = SubjectForm(request.POST,instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect('add_subject')
    else:
        form = SubjectForm(instance=post)
    return render(request, 'edit_subject.html', {'form': form})
def delete_subject(request, pk):
    subject = Subject.objects.get(subject_id=pk)
    subject.delete()
    return redirect("add_subject")
def assign_subjects_to_teachers(request):
    school_id = request.session['schoolname']
    academic_year = request.session['academic_year']
    teacher = StaffDetail.objects.filter(school_id=school_id)
    class_id = Class.objects.filter(school_id=school_id,academic_year=academic_year)
    if request.method == "POST":
        form = AssignTeacherForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect("assign_teachers_view")  
    else:
        form = AssignTeacherForm()
    return render(request, 'assign_subjects_to_teachers.html', {'form':form,'teacher':teacher,'class_id':class_id})
def edit_assign_subjects_to_teachers(request,pk):
    post = get_object_or_404(AssignSubjectTeacher, pk=pk)
    teacher = AssignSubjectTeacher.objects.filter(assign_subject_teacher_id=pk)
    if request.method == "POST":
        form = AssignTeacherEditForm(request.POST,instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect("assign_teachers_view")
    else:
        form = AssignTeacherEditForm(instance=post)
    return render(request, 'edit_assign_subjects_to_teachers.html', {'form':form,'teacher':teacher})
def delete_assign_teacher(request,pk):
    subject = AssignSubjectTeacher.objects.get(assign_subject_teacher_id=pk)
    subject.delete()
    return redirect("assign_teachers_view")
def add_exam(request):
    school_id = request.session['schoolname']
    academic_year = request.session['academic_year']
    exam = Exam.objects.filter(school_id=school_id ,class_id__in=Class.objects.filter(academic_year=academic_year)).select_related('class_id').select_related('section_id').order_by('exam_id')
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(exam,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    if request.method == "POST":
        form = ExamForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            messages.success(request, 'Exam Added Successfully')
            form = ExamForm()
    else:
        form = ExamForm()
    return render(request, 'add_exam.html', {'form':form,'exam':users})
def edit_exam(request,pk):
    post = get_object_or_404(Exam, pk=pk)
    if request.method == "POST":
        form = ExamEditForm(request.POST,instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect('add_exam')
    else:
        form = ExamEditForm(instance=post)
    return render(request, 'edit_exam.html', {'form': form})
def delete_exam(request, pk):
    exam = Exam.objects.get(exam_id=pk)
    exam.delete()
    return redirect("add_exam")
def exams(request):
    return render(request, 'exams.html', {})
def change_password(request):
    return render(request, 'change_password.html', {})
def pagination_count_year(request):
    school_id = request.session['schoolname']
    page_count = SchoolSetting.objects.filter(school_id=school_id)
    if request.method == "POST":
        settings_id = request.POST.get('settings_id')
        count = request.POST.get('page_count')
        cnt = int(count)
        academic_year = request.POST.get('academic_year')
        set_id = int(settings_id)
        SchoolSetting.objects.filter(settings_id=set_id).update(academic_year=academic_year,pagination_count=cnt)
        messages.success(request,"Settings Updated Successfully")
    return render(request, 'other_settings.html', {'page_count':page_count})
def pagination_count(request):
    school_id = request.session['schoolname']
    page_count = SchoolSetting.objects.filter(school_id=school_id)
    if request.method == "POST":
        settings_id = request.POST.get('settings_id')
        count = request.POST.get('page_count')
        cnt = int(count)
        academic_year = request.POST.get('academic_year')
        set_id = int(settings_id)
        SchoolSetting.objects.filter(settings_id=set_id).update(academic_year=academic_year,pagination_count=cnt)
        messages.success(request,"Settings Updated Successfully")
    return render(request, 'admin_settings.html', {'page_count':page_count})
def student_section(request):
    school_id=request.session['schoolname']
    academic_year = request.session['academic_year']
    class_name = Class.objects.filter(school_id=school_id,academic_year=academic_year)
    student = StudentDetail.objects.filter(school_id=school_id)
    
    if "GET" == request.method:
        return render(request, 'student_section.html', {'student':student,'cls_name':class_name})
    else:
        file = request.FILES['excel_file']
        school_id = request.session['schoolname']
        ids = SchoolDetail.objects.get(id=int(school_id))
        workbook = openpyxl.load_workbook(file, read_only=True)
        # Get name of the first sheet and then open sheet by name
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)

        data = []
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
            register_num = row[0].value
            student_name = row[1].value
            cls_name = row[2].value
            sec_name = row[3].value
            get_student_id = StudentDetail.objects.only('id').get(register_number=register_num).id
            stud_id = StudentDetail.objects.get(id=int(get_student_id))
            class_id = Class.objects.only('class_id').get(class_name=cls_name,school_id=school_id,academic_year=academic_year).class_id
            student_class = Class.objects.get(class_id=int(class_id))
            se_id = Section.objects.filter(section_name=sec_name)
            section_id = Section.objects.only('section_id').get(section_name=sec_name,class_id=student_class).section_id
            student_section = Section.objects.get(section_id=int(section_id))
            user_exist = StudentSection.objects.filter(student_id=stud_id,academic_year=academic_year)

            if user_exist:
                messages.success(request, student_name + " Class Already Assigned")
            else:
                student = StudentSection()
                student.student_id = stud_id
                student.class_id = student_class
                student.section_id = student_section
                student.academic_year = academic_year
                data.append(student)
        # Bulk create data
        StudentSection.objects.bulk_create(data)
        return redirect("view_student_section")

    return render(request, 'student_section.html', {'student':student,'cls_name':class_name})
def edit_student_section(request,pk):
    post = get_object_or_404(StudentSection, pk=pk)
    if request.method == "POST":
        form = StudentSectionEditForm(request.POST,instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            post.save()
            return redirect("view_student_section")
    else:
        form = StudentSectionEditForm(instance=post)
    return render(request, 'edit_student_section.html', {'form': form , 'post' :post})

def delete_student_section(request, pk):
    student = StudentSection.objects.get(student_section_id=pk)
    student.delete()
    return redirect("view_student_section")

def student_result(request):
    return render(request, 'student_result.html', {})

def student_mark(request):
    return render(request, 'student_mark.html', {})

def ajax_exam_class_search(request):
    academic_year = request.POST.get('academic_year')
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    post = '''SELECT school_class.class_id,school_class.class_name from school_class where school_class.academic_year='%s' AND school_class.school_id_id='%d' ''' % (academic_year,school_id)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data = {
    'class_name':row
    }
    return JsonResponse(data)

def student_view_marks(request):
    student_id = request.session['student_id']
    student_name = StudentSection.objects.filter(student_id=student_id)
    return render(request, 'student_view_marks.html', {'student_name':student_name})

def student_view_diary_notes(request):
    student_id = request.session['student_id']
    cursor = connection.cursor()
    student = ''' SELECT * from school_studentsection where school_studentsection.student_id_id='%d' ''' % (student_id)
    query = cursor.execute(student)
    row = cursor.fetchone()
    cls_id = row[2]
    sec_id = row[3]
    task = StudentDiary.objects.filter(class_id=cls_id,section_id=sec_id)
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(task,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    return render(request, 'student_view_diary_notes.html', {'student':users})

def manage_student_marks(request):
    school_id=request.session['schoolname']
    academic_year=request.session['academic_year']
    staff_id=request.session['user_id']
    staff = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,class_id__in=Class.objects.filter(academic_year=academic_year,school_id_id=school_id)).select_related('class_id').select_related('section_id').select_related('subject_id')
    mark=Exam.objects.filter(school_id_id=school_id,class_id__in=Class.objects.filter(academic_year=academic_year)).select_related('class_id')   
    class_id = staff.values('class_id')
    section_id = staff.values('section_id')

    student = StudentSection.objects.filter(class_id__in=AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,class_id__in=Class.objects.filter(
    academic_year=academic_year,school_id_id=school_id)).select_related('class_id').select_related('section_id').values('class_id'),section_id__in=AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,class_id__in=Class.objects.filter(
    academic_year=academic_year,school_id_id=school_id)).select_related('class_id').select_related('section_id').values('section_id'))

    return render(request, 'manage_student_marks.html', {'mark':mark,'staff':staff,'class_id':class_id,'section_id':section_id,'student':student})
def view_manage_student_marks(request):
    if request.session['user_id']:
        cls_id = request.GET.get('cls_id')
        sec_id = request.GET.get('sec_id')
        class_id = int(cls_id)
        section_id = int(sec_id)
        school_id = request.session['schoolname']
        staff_id = request.session['user_id']
        academic_year = request.session['academic_year']
        student = StudentSection.objects.filter(class_id_id=class_id,
        section_id_id=section_id,academic_year=academic_year,student_id__in=StudentDetail.objects.filter(school_id=school_id)) 
        return render(request, 'view_manage_student_marks.html', {'post':student})
    else:
        return redirect("teacher_login")

def add_student_diary_notes(request):
    if request.GET.get('class_id') and request.GET.get('sec_id'):
        class_id = request.GET.get('class_id')
        cls_id = int(class_id)
        section_id = request.GET.get('sec_id')
        sec_id = int(section_id)
        staff_id=request.session['user_id']
        if request.method == "POST":
            homework_classwork = int(request.POST.get('homework_classwork'))
            subject = request.POST.get('subject_id_id')
            assigned_date = request.POST.get('assigned_date')
            assigned_date =datetime.strptime(assigned_date, "%d-%m-%Y").strftime('%Y-%m-%d')
            diary_notes = request.POST.get('diary_notes')
            StudentDiary.objects.create(subject_id_id=subject,assigned_date=assigned_date, diary_note=diary_notes,class_id_id=cls_id,section_id_id=sec_id,staff_id_id=staff_id, homework_classwork=homework_classwork)
            messages.success(request,'Diary Added Successfully')
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        student = StudentSection.objects.filter(class_id=cls_id,academic_year=academic_year,section_id=sec_id,student_id_id__in=StudentDetail.objects.filter(school_id=school_id) )
        subject = AssignSubjectTeacher.objects.filter(section_id=sec_id,class_id__in=Class.objects.filter(academic_year=academic_year,class_id=int(cls_id)), staff_id=staff_id)
        return render(request,'add_student_diary_notes.html',{'subject':subject,'student':student})
    else:
        return redirect("teacher_class_diary")

def view_student_diary_notes(request):
    diary_data = dict()
    if request.method == "POST":
        diary_id = request.POST.get('diary_id')
        ids = int(diary_id)
        diary = StudentDiary.objects.get(diary_id=ids)
        diary.delete()
    if request.GET.get('class_id') and request.GET.get('sec_id'):
        class_id = request.GET.get('class_id')
        cls_id = int(class_id)
        section_id = request.GET.get('sec_id')
        sec_id = int(section_id)
        staff_id=request.session['user_id']
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        diary_data = StudentDiary.objects.filter(class_id_id__in=Class.objects.filter(class_id=cls_id,academic_year=academic_year),section_id_id=sec_id,staff_id_id=staff_id )
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(diary_data,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
    return render(request,'view_student_diary_notes.html',{'diary_data':users})    
def view_student_report(request):
    return render(request, 'view_student_report.html', {})
def teacher_change_password(request):
    staff_id=request.session['user_id']
    get_teacher=""
    if request.method == "POST":
        get_psw =  request.POST['password']
        password = StaffDetail.objects.filter(staff_id=staff_id).update(password=get_psw)
        messages.success(request, 'Password Updated Successfully')
    else:
        get_teacher = StaffDetail.objects.filter(staff_id=staff_id)
    return render(request,'teacher_change_password.html',{"teacher": get_teacher})
def teacher_other_settings(request):
    return render(request, 'teacher_other_settings.html', {})

def view_student_section(request):
    academic_year = request.session['academic_year'];
    school_id = request.session['schoolname'];
    row = StudentSection.objects.filter(academic_year=academic_year,student_id__in=StudentDetail.objects.filter(school_id_id=school_id)).select_related('student_id').select_related('class_id').select_related('section_id').order_by('-student_section_id')
    length = len(row)
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(row,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    return render(request, 'view_student_section.html', {'row' : users,'length':length})

def select_section(request):
    school_id = request.session['schoolname']
    class_id = request.POST.get('class_id')
    academic_year = request.session['academic_year']
    cursor = connection.cursor()
    post = '''SELECT * from school_class INNER JOIN school_section ON
    school_class.class_id=school_section.class_id_id where school_section.class_id_id='%d'
    AND school_class.school_id_id='%d' AND school_class.academic_year='%s' AND school_section.school_id_id='%d' ''' % (int(class_id),int(school_id),academic_year,int(school_id))
    query = cursor.execute(post)
    row = cursor.fetchall()
    data =  {

        'rowval':row
    }
    return JsonResponse(data)

def select_teacher_section(request):
    school_id = request.session['schoolname']
    class_id = request.POST.get('class_id')
    academic_year = request.session['academic_year']
    teacher_id = request.session['user_id']
    cursor = connection.cursor()
    post = '''SELECT COUNT(school_class.class_id),COUNT(school_section.section_id), from school_class INNER JOIN school_section ON
    school_class.class_id=school_section.class_id_id INNER JOIN school_assignsubjectteacher ON
    school_assignsubjectteacher.section_id_id=school_section.section_id  where school_section.class_id_id='%d'
    AND school_class.school_id_id='%d' AND school_class.academic_year='%s' AND school_section.school_id_id='%d'
    AND school_assignsubjectteacher.staff_id_id='%d' GROUP BY  school_class.class_id''' % (int(class_id),int(school_id),academic_year,int(school_id),int(teacher_id))
    query = cursor.execute(post)
    row = cursor.fetchall()
    data =  {

        'rowval':row
    }
    return JsonResponse(data)

def select_academic_year(request):
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    post = '''SELECT * from school_schoolsetting where school_schoolsetting.school_id_id='%d' ''' %(int(school_id))
    query = cursor.execute(post)
    row = cursor.fetchone()
    year = row[1]
    request.session['academic_year'] = year
    academic_year =  request.session['academic_year']
    data = {

    'academic_year':academic_year

    }
    return JsonResponse(data)

def select_class(request):
    school_id = request.POST.get('school_id')
    academic_year = request.session['academic_year']
    cursor = connection.cursor()
    post = '''SELECT * from school_class where school_class.academic_year='%s' ''' % (academic_year)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data =  {

        'row':row
    }
    return JsonResponse(data)

def student_section_class_select(request):
    school_id = request.session['schoolname']
    class_id = request.POST.get('class_id')
    academic_year = request.session['academic_year']

    cursor = connection.cursor()
    post = '''SELECT * from school_class INNER JOIN school_section ON
    school_class.class_id=school_section.class_id_id where school_section.class_id_id='%d'
    AND school_class.school_id_id='%d' AND school_class.academic_year='%s' ''' % (int(class_id),int(school_id),academic_year)
    query = cursor.execute(post)
    row = cursor.fetchall()

    data = {
        'row':row
    }

    return JsonResponse(data)

def section_selection_student(request):

    sec_id = request.GET.get('id')
    cursor = connection.cursor()
    post = '''SELECT * from school_StudentSection  where school_StudentSection.student_section_id='%d'
     ''' % (int(sec_id))
    query = cursor.execute(post)
    row = cursor.fetchall()

    data = {
        'row':row
    }
    return JsonResponse(data)

def select_subject(request):
    ids = request.POST.get('class_id')
    class_id = int(ids)
    sec_id = request.POST.get('section_id')
    section_id = int(sec_id)
    academic_year = request.session['academic_year']
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    post = '''SELECT * from school_class INNER JOIN school_section ON school_class.class_id=school_section.class_id_id
    INNER JOIN school_subject ON school_subject.section_id_id=school_section.section_id where school_subject.class_id_id ='%d'
    AND school_subject.section_id_id='%d' AND school_class.school_id_id='%d' AND school_class.academic_year='%s' AND school_section.school_id_id='%d' ''' % (class_id,section_id,school_id,academic_year,school_id)
    query = cursor.execute(post)
    row = cursor.fetchall()

    data = {
        'subject':row
    }
    return JsonResponse(data)

def assign_teachers_view(request):
    academic_year = request.session['academic_year']
    school_id = request.session['schoolname']
    teachers = AssignSubjectTeacher.objects.filter(academic_year=academic_year,staff_id__in=StaffDetail.objects.filter(school_id_id=school_id)).select_related('class_id').select_related('section_id').select_related('subject_id').select_related('staff_id').order_by('-assign_subject_teacher_id')
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(teachers,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    return render(request,'assign_teachers_view.html',{'teachers':users})

def admin_change_password(request):
    admin_id = request.session['schoolname']
    get_admin=""
    if request.method == "POST":
        get_psw =  request.POST['password']
        password = SchoolDetail.objects.filter(id=admin_id).update(password=get_psw)
        messages.success(request, 'Password Updated Successfully')
    else:
        get_admin = SchoolDetail.objects.filter(id=admin_id)
    return render(request,'change_password.html',{"school": get_admin})

def manage_students_marks(request):
    ex=""
    school_id=request.session['schoolname']
    academic_year=request.session['academic_year']
    staff_id=request.session['user_id']
    query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year)
    staff = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('class_id')).order_by('class_id__class_name')
    class_id = query.values('class_id')
    section_id = query.values('section_id')
    count = len(class_id)
    for i in range(0,count):
    	exams = Exam.objects.filter(class_id__in=class_id,section_id__in=section_id)
    	ex = exams.values('class_id', 'section_id','class_id__class_name','section_id__section_name','exam_id','exams').annotate(count = Count('class_id'))
    
    return render(request,'manage_students_marks.html',{'staff':staff,'class_id':class_id,'exams':ex,'section_id':section_id,'query':query})
def is_class_teacher(request):
    school_id=request.session['schoolname']
    academic_year=request.session['academic_year']
    staff_id=request.session['user_id']
    query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year,is_class_teacher='yes')
    staff = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('class_id')).order_by('class_id__class_name')
    return render(request,'is_class_teacher.html',{'staff':staff})
def class_student_mark_details(request,class_id,section_id,exam_id):
    cursor = connection.cursor()
    post = '''SELECT school_mark.student_reg_no, SUM(school_mark.mark), Avg(school_mark.mark) from school_mark INNER JOIN school_studentdetail ON school_mark.student_reg_no=school_studentdetail.register_number
    where school_mark.class_id ='%d'AND school_mark.section_id='%d' AND school_mark.exam_id='%d' GROUP BY school_mark.student_reg_no order by school_mark.student_reg_no''' % (class_id,section_id,exam_id)
    ex = cursor.execute(post)
    row = cursor.fetchall()
    sub1 = '''SELECT  COUNT(school_subject.subject_id),school_subject.subject_name,school_subject.subject_id from school_mark INNER JOIN school_subject ON school_mark.subject_id=school_subject.subject_id
    where school_mark.class_id ='%d'AND school_mark.section_id='%d'  GROUP BY school_subject.subject_id order by school_subject.subject_id''' % (class_id,section_id)
    ex1 = cursor.execute(sub1)
    sub_name = cursor.fetchall()
    cnt  = len(sub_name)
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(row,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    return render(request,'class_student_mark_details.html',{'row':users,'class_id':class_id,
    'section_id':section_id,'exam_id':exam_id,'sub_name':sub_name,'tot':cnt})
def select_class_exam(request):
    if request.GET.get('cls_id') and request.GET.get('sec_id'):
        cls_id = request.GET.get('cls_id')
        sec_id = request.GET.get('sec_id')
        class_id = int(cls_id)
        section_id = int(sec_id)
        school_id = request.session['schoolname']
        exams = Exam.objects.filter(class_id_id=class_id,section_id_id=section_id,school_id_id=school_id)
        return render(request,'select_class_exam.html',{'exams':exams})
    else:
        return redirect("is_class_teacher")
def ajax_student_marks(request):
    cls_id = request.POST.get('class_id')
    sec_id = request.POST.get('section_id')
    sub_id = request.POST.get('subject_id')
    request.session['class_id'] = cls_id
    data = {
    'cls_id':cls_id,
    'sec_id':sec_id,
    'sub_id':sub_id
    }
    return JsonResponse(data)
def add_student_marks(request,cls_id,sec_id,exam_id):
  
    rc = dict()
    st = StudentSection.objects.filter(class_id=cls_id,section_id=sec_id)
    cnt = len(st)
    teacher_id = request.session['user_id']
    cursor = connection.cursor()
    get_mark = ''' SELECT school_mark.*,school_studentdetail.register_number,school_studentdetail.student_name,school_subject.subject_name,school_class.class_name,school_section.section_name from school_mark
                    LEFT JOIN school_studentdetail ON school_studentdetail.register_number=school_mark.student_reg_no
                    LEFT JOIN school_subject ON school_subject.subject_id=school_mark.subject_id
                    LEFT JOIN school_class ON school_class.class_id=school_mark.class_id
                    LEFT JOIN school_section ON school_section.section_id=school_mark.section_id
                    WHERE school_mark.class_id=%d AND school_mark.section_id=%d AND school_mark.staff_id=%d AND school_mark.exam_id=%d '''%(int(cls_id),int(sec_id),int(teacher_id),int(exam_id))
    sub = cursor.execute(get_mark)
    mark = cursor.fetchall()
    school_id = request.session['schoolname']
    academic_year = request.session['academic_year']

    if request.method == "POST":
        if request.POST.getlist('mark_id[]'):
            ids = request.POST.getlist('mark_id[]')
            mark = request.POST.getlist('student_mark[]')
            length=len(mark)
            for row in range(0,length):
                if mark[row]=="" :
                    mark[row]=0;
                Mark.objects.filter(mark_id=ids[row]).update(mark=mark[row])
            return redirect("manage_students_marks")
        else :
            subject_id = request.POST.getlist('subject_id[]')
            mark = request.POST.getlist('student_mark[]')
            student_id = request.POST.getlist('stud_reg_no[]')
            staff_id = request.POST.getlist('staff_id[]')
            length=len(mark)
            for row in range(0,length):
                if mark[row]=="":
                    mark[row]=0;
                Mark.objects.create(student_reg_no=student_id[row],class_id=cls_id,section_id=sec_id,subject_id=subject_id[row],staff_id=staff_id[row],mark=mark[row],exam_id=exam_id)
            return redirect("manage_students_marks")
    #exam = Exam.objects.filter(section_id=int(sec_id),school_id=school_id,class_id__in=Class.objects.filter(academic_year=academic_year,class_id=int(cls_id)))
    exam = Exam.objects.filter(exam_id=exam_id)
    student = StudentSection.objects.filter(class_id=cls_id,academic_year=academic_year,section_id=sec_id,student_id_id__in=StudentDetail.objects.filter(school_id=school_id))
    subject = AssignSubjectTeacher.objects.filter(section_id=sec_id,class_id__in=Class.objects.filter(academic_year=academic_year,class_id=int(cls_id)),staff_id=teacher_id)
        
    return render(request,'add_student_marks.html',{'exam':exam,'student':student,'subject':subject,'student_reg':rc, 'mark' :mark,'cnt':cnt})

def import_mark(request,class_id,section_id,exam_id):
    st = StudentSection.objects.filter(class_id=class_id,section_id=section_id)
    cnt = len(st)

    staff_id = request.session['user_id']
    clss_id = Section.objects.filter(section_id=section_id).select_related('class_id')
    val = clss_id.values('class_id_id','section_id')
    ex_name = Exam.objects.filter(exam_id=exam_id)
    subject_name = AssignSubjectTeacher.objects.filter(class_id=class_id,section_id=section_id,staff_id=staff_id).order_by('subject_id')
    if "GET" == request.method:
        return render(request, 'import_mark.html', {'subject_name':subject_name,'ex_name':ex_name,'val':val,'cnt':cnt})
    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)

        worksheet = wb["Mark Sheet"]

        mylist = []
        for cell in worksheet['C']:
            if cell.value != None:
                mylist.append(cell.value)

        mark1 = []
        for cell in worksheet['D']:
            if cell.value != None:
                mark1.append(cell.value)
        m1 = len(mark1)

        mark2 = []
        for cell in worksheet['E']:
            if cell.value != None:
                mark2.append(cell.value)
        m2 = len(mark2)

        mark3 = []
        for cell in worksheet['F']:
            if cell.value != None:
                mark3.append(cell.value)
        m3 = len(mark3)

        mark4 = []
        for cell in worksheet['G']:
            if cell.value != None:
                mark4.append(cell.value)
        m4 = len(mark4)

        mark5 = []
        for cell in worksheet['H']:
            if cell.value != None:
                mark5.append(cell.value)
        m5 = len(mark5)

        mark6 = []
        for cell in worksheet['I']:
            if cell.value != None:
                mark6.append(cell.value)
        m6 = len(mark6)

        mark7 = []
        for cell in worksheet['J']:
            if cell.value != None:
                mark7.append(cell.value)
        m7 = len(mark7)

        cls_id = class_id
        sec_id = section_id
        ex_id = exam_id

        staff_id = request.session['user_id']
        sub_id = request.POST.getlist('sub_id[]')
        lenthsub = len(sub_id)
        
        if lenthsub == 1:
            sub1 = sub_id[0]
        if lenthsub == 2:
            sub1 = sub_id[0]
            sub2 = sub_id[1]
        if lenthsub == 3:
            sub1 = sub_id[0]
            sub2 = sub_id[1]
            sub3 = sub_id[2]
        if lenthsub == 4:
            sub1 = sub_id[0]
            sub2 = sub_id[1]
            sub3 = sub_id[2]
            sub4 = sub_id[3]
        if lenthsub == 5:
            sub1 = sub_id[0]
            sub2 = sub_id[1]
            sub3 = sub_id[2]
            sub4 = sub_id[3]
            sub5 = sub_id[4]
        if lenthsub == 6:
            sub1 = sub_id[0]
            sub2 = sub_id[1]
            sub3 = sub_id[2]
            sub4 = sub_id[3]
            sub5 = sub_id[4]
            sub6 = sub_id[5]
        if lenthsub == 7:
            sub1 = sub_id[0]
            sub2 = sub_id[1]
            sub3 = sub_id[2]
            sub4 = sub_id[3]
            sub5 = sub_id[4]
            sub6 = sub_id[5]
            sub7 = sub_id[6]

        length=len(mylist)

        for row in range(1,length):
            if m1 > 0 and sub1!="":
                Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub1,mark=mark1[row])
            if m2 > 0 and sub2!="":
                Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub2,mark=mark2[row])
            if m3 > 0 and sub3!="":
                Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub3,mark=mark3[row])
            if m4 > 0 and sub4!="":
                Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub4,mark=mark4[row])
            if m5 > 0 and sub5!="":
                Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub5,mark=mark5[row])
            if m6 > 0 and sub6!="":
                Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub6,mark=mark6[row])
            if m7 > 0 and sub7!="":
                Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub7,mark=mark7[row])
        
        messages.success(request, "Mark Added Successfully")

        return render(request, 'import_mark.html', {'subject_name':subject_name,'ex_name':ex_name,'val':val,'cnt':cnt})

def overall_subject_chart(request):
    cls_id = request.GET.get('cls_id')
    sec_id = request.GET.get('sec_id')
    class_id = int(cls_id)
    section_id = int(sec_id)
    school_id = request.session['schoolname']
    staff_id = request.session['user_id']
    academic_year = request.session['academic_year']
    query = StudentSection.objects.filter(class_id_id=class_id,section_id_id=section_id,
    academic_year=academic_year)
    student_count = query.values('student_id_id').aggregate(count = Count('student_id_id'))
    cursor = connection.cursor()
    subject = '''SELECT COUNT(school_mark.subject_id),SUM(school_mark.mark),COUNT(school_subject.subject_name) from school_class INNER JOIN school_mark ON school_class.class_id=school_mark.class_id
    INNER JOIN school_subject ON school_subject.subject_id=school_mark.subject_id
    where school_class.academic_year='%s' AND school_mark.class_id='%d' AND school_mark.section_id='%d' 
    AND school_mark.staff_id='%d' AND school_class.school_id_id='%d' GROUP BY school_mark.subject_id''' % (academic_year,class_id,section_id,int(staff_id),int(school_id))
    sub = cursor.execute(subject)
    row = cursor.fetchall()

    return render(request,'overall_subject_chart.html',{'student_count':student_count,'row':row,'query':query})

def ajax_student_grap(request):
    cls_id = request.POST.get('cls_id')
    sec_id = request.POST.get('sec_id')
    class_id = int(cls_id)
    section_id = int(sec_id)
    school_id = request.session['schoolname']
    staff_id = request.session['user_id']
    academic_year = request.session['academic_year']
    query = StudentSection.objects.filter(class_id_id=class_id,section_id_id=section_id,
    academic_year=academic_year)
    student_count = query.values('student_id_id').aggregate(count = Count('student_id_id'))
    cursor = connection.cursor()
    subject = '''SELECT COUNT(school_mark.subject_id),SUM(school_mark.mark),school_subject.subject_name,COUNT(school_mark.student_reg_no),school_subject.subject_id from school_class INNER JOIN school_mark ON school_class.class_id=school_mark.class_id
    INNER JOIN school_subject ON school_subject.subject_id=school_mark.subject_id
    where school_class.academic_year='%s' AND school_mark.class_id='%d' AND school_mark.section_id='%d' 
    AND school_mark.staff_id='%d' AND school_class.school_id_id='%d' GROUP BY school_subject.subject_id''' % (academic_year,class_id,section_id,int(staff_id),int(school_id))
    sub = cursor.execute(subject)
    row = cursor.fetchall()

    data = {
    'row':row,
    'student_count':student_count
    }
    return JsonResponse(data)

def single_student_subject_marks_chart(request,pk):
	school_id=request.session['schoolname']
	student_name = StudentDetail.objects.filter(register_number=pk)
	cursor = connection.cursor()
	post = ''' SELECT SUM(school_mark.mark),school_mark.subject_id
	from school_mark INNER JOIN school_subject ON 
	school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' GROUP BY school_mark.subject_id ''' % (pk)
	sub = cursor.execute(post)
	row = cursor.fetchall()
	lenth = len(row)

	subject_name = '''SELECT COUNT(school_subject.subject_id),school_subject.subject_name,school_subject.subject_id from school_mark INNER JOIN school_subject ON 
	school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' 
	GROUP BY school_subject.subject_name,school_subject.subject_id ''' % (pk)
	query = cursor.execute(subject_name)
	sub_name =cursor.fetchall()

	return render(request,'single_student_subject_marks_chart.html',{'mark':row,'student_name':student_name,'lenth':lenth,'subject_name':sub_name})
def student_subject_marks_chart(request,pk):
    student_name = StudentDetail.objects.filter(register_number=pk)
    cursor = connection.cursor()
    mark = ''' SELECT SUM(school_mark.mark),school_mark.subject_id
    from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' GROUP BY school_mark.subject_id ''' % (pk)
    sub = cursor.execute(mark)
    row = cursor.fetchall()
    lenth = len(row)
    subject_name = '''SELECT COUNT(school_subject.subject_id),school_subject.subject_name,school_subject.subject_id from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' 
    GROUP BY school_subject.subject_name,school_subject.subject_id ''' % (pk)
    query = cursor.execute(subject_name)
    sub_name =cursor.fetchall()
    return render(request,'student_subject_marks_chart.html',{'mark':row,'student_name':student_name,'lenth':lenth,'subject_name':sub_name})
def edit_students_mark(request,pk):
    staff_id=request.session['user_id']
    academic_year = request.session['academic_year']
    cursor = connection.cursor()
    mark = '''SELECT * from school_mark INNER JOIN school_class ON school_mark.class_id=school_class.class_id
    where school_mark.student_reg_no='%s' AND school_mark.staff_id='%d' 
    AND school_class.academic_year='%s' ''' % (pk,int(staff_id),academic_year)
    query = cursor.execute(mark)
    row = cursor.fetchall()
    return render(request,'edit_students_mark.html',{'row':row})

def ajax_subject_count(request):
    student_reg_no = request.GET.get('ids')
    cursor = connection.cursor()
    post = ''' SELECT SUM(school_subject.subject_id),school_subject.subject_name,school_subject.subject_id
    from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' GROUP BY school_subject.subject_id''' % (student_reg_no)
    sub = cursor.execute(post)
    row = cursor.fetchall()

    post2 = ''' SELECT COUNT(school_mark.subject_id) from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' GROUP BY school_mark.subject_id''' % (student_reg_no)
    sub1 = cursor.execute(post2)
    row1 = cursor.fetchall()

    lenth = len(row)
    data = {
    'lenth':lenth,
    'row':row,
    'row1':row1
    }
    return JsonResponse(data)
def teacher_class_diary(request):
    school_id=request.session['schoolname']
    academic_year=request.session['academic_year']
    staff_id=request.session['user_id']
    query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year).select_related('class').select_related('section')
    query = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('section_id'), count2= Count('class_id')).order_by('class_id__class_name')
    return render(request,'teacher_class_diary.html',{'staff':query})

def manage_exam_mark_chart(request):
    school_id=request.session['schoolname']
    academic_year=request.session['academic_year']
    staff_id=request.session['user_id']
    query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year)
    query = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('class_id')).order_by('class_id__class_name')
    return render(request,'manage_exam_mark_chart.html',{'staff':query})

def particular_exam_mark_chart(request,pk,cls_id,sec_id):
    cursor = connection.cursor()
    post = ''' SELECT school_subject.subject_id, school_subject.subject_name  from school_subject INNER JOIN school_mark  
    ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    GROUP BY school_subject.subject_id order by school_subject.subject_id'''  % (cls_id,sec_id,pk)
    query = cursor.execute(post)
    row = cursor.fetchall()

    post1 = ''' SELECT school_mark.mark, school_mark.subject_id,school_mark.exam_id from school_subject INNER JOIN school_mark  
    ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    order by school_exam.exam_id,school_subject.subject_id'''  % (cls_id,sec_id,pk)
    query1 = cursor.execute(post1)
    row1 = cursor.fetchall()

    post2 = ''' SELECT COUNT(school_exam.exam_id),school_exam.exams, SUM(school_mark.mark) from school_mark INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    Group By school_exam.exam_id order by school_exam.exam_id '''  % (cls_id,sec_id,pk)
    query2 = cursor.execute(post2)
    row2 = cursor.fetchall()

    post5 = ''' SELECT COUNT(school_mark.exam_id) from school_mark 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    Group By school_mark.exam_id '''  % (cls_id,sec_id,pk)
    query5 = cursor.execute(post5)
    row5 = cursor.fetchall()
    count3 = len(row5)

    exam_id = Exam.objects.filter(class_id=cls_id,section_id=sec_id).order_by('exam_id')
    exam = Mark.objects.filter(student_reg_no=pk)
    cnt = exam.values('exam_id','student_reg_no').annotate(count=Count('exam_id'),count1=Count('student_reg_no'))

    return render(request,'particular_exam_mark_chart.html',{'row':row,'length':cnt,'exam_id':exam_id,'row1':row1,'row2':row2,'pk':pk,
    'len_exam':count3})
def get_exam_name(request):
    cls_id = request.POST.get('cls_id')
    sec_id = request.POST.get('sec_id')
    pk = request.POST.get('student_id')
    class_id = int(cls_id)
    section_id = int(sec_id)
    cursor = connection.cursor()
    post = ''' SELECT school_mark.mark, school_mark.subject_id,school_mark.exam_id, school_exam.exams from school_subject INNER JOIN school_mark  
    ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    order by school_exam.exam_id,school_subject.subject_id'''  % (class_id,section_id,pk)
    query = cursor.execute(post)
    row = cursor.fetchall()

    data = {
    'row':row
    }
    return JsonResponse(data)
def Student_particular_exam_mark_chart(request,pk,cls_id,sec_id):
    cursor = connection.cursor()
    post = ''' SELECT school_subject.subject_id, school_subject.subject_name  from school_subject INNER JOIN school_mark  
    ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    GROUP BY school_subject.subject_id order by school_subject.subject_id'''  % (cls_id,sec_id,pk)
    query = cursor.execute(post)
    row = cursor.fetchall()

    post1 = ''' SELECT school_mark.mark, school_mark.subject_id,school_mark.exam_id from school_subject INNER JOIN school_mark  
    ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    order by school_exam.exam_id,school_subject.subject_id'''  % (cls_id,sec_id,pk)
    query1 = cursor.execute(post1)
    row1 = cursor.fetchall()

    post2 = ''' SELECT COUNT(school_exam.exam_id),school_exam.exams, SUM(school_mark.mark) from school_mark INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    Group By school_exam.exam_id order by school_exam.exam_id '''  % (cls_id,sec_id,pk)
    query2 = cursor.execute(post2)
    row2 = cursor.fetchall()

    post5 = ''' SELECT COUNT(school_mark.exam_id) from school_mark 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    Group By school_mark.exam_id '''  % (cls_id,sec_id,pk)
    query5 = cursor.execute(post5)
    row5 = cursor.fetchall()
    count3 = len(row5)

    exam_id = Exam.objects.filter(class_id=cls_id,section_id=sec_id).order_by('exam_id')
    exam = Mark.objects.filter(student_reg_no=pk)
    cnt = exam.values('exam_id','student_reg_no').annotate(count=Count('exam_id'),count1=Count('student_reg_no'))

    return render(request,'Student_particular_exam_mark_chart.html',{'row':row,'length':cnt,'exam_id':exam_id,'row1':row1,'row2':row2,'pk':pk,
    'len_exam':count3})
def select_exam(request):
    cls_id = request.POST.get('class_id')
    class_id = int(cls_id)
    sec_id = request.POST.get('section_id')
    section_id = int(sec_id)
    staff_id = request.session['user_id']
    cursor = connection.cursor()
    sql= ''' SELECT * from school_exam where school_exam.class_id_id='%d' AND school_exam.section_id_id='%d'
    ''' % (class_id,section_id)
    res = cursor.execute(sql)
    exam_name = cursor.fetchall()
    post = ''' SELECT COUNT(school_mark.exam_id) ,COUNT(school_mark.subject_id),SUM(school_mark.mark),COUNT(school_mark.student_reg_no) from school_mark  where school_mark.class_id='%d' AND school_mark.section_id='%d'
    AND school_mark.staff_id='%d' GROUP BY school_mark.subject_id''' % (class_id,section_id,staff_id)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data = {
    'row':row,
    'exam_name':exam_name
    }
    return JsonResponse(data)

def select_school_exam(request):
    cls_id = request.POST.get('class_id')
    class_id = int(cls_id)
    sec_id = request.POST.get('section_id')
    section_id = int(sec_id)
    cursor = connection.cursor()
    sql= ''' SELECT  school_exam.class_id_id,school_exam.section_id_id,school_exam.exam_id,school_exam.exams from school_exam where school_exam.class_id_id='%d' AND school_exam.section_id_id='%d'
    ''' % (class_id,section_id)
    res = cursor.execute(sql)
    exam_name = cursor.fetchall()
    data = {
    'exam_name':exam_name
    }
    return JsonResponse(data)

def select_school_details(request):
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    sql= ''' SELECT * from school_schooldetail where school_schooldetail.id='%d' ''' % (school_id)
    res = cursor.execute(sql)
    school_details = cursor.fetchall()
    data ={
        'school_details':school_details
    }
    return JsonResponse(data)

def search_student_marks_ajax(request):
    cls_id = request.POST.get('class_id')
    class_id = int(cls_id)
    sec_id = request.POST.get('section_id')
    section_id = int(sec_id)
    ex_id = request.POST.get('exam_id')
    exam_id = int(ex_id)
    student_reg_no = request.POST.get('student_reg_no')
    cursor = connection.cursor()
    stud = '''SELECT * from school_studentdetail'''
    result = cursor.execute(stud)
    student_id = cursor.fetchall()

    if  student_reg_no == "":
        sql= ''' SELECT COUNT(school_mark.student_reg_no),SUM(school_mark.mark),school_mark.student_reg_no,COUNT(school_mark.subject_id) from school_mark where school_mark.class_id='%d' AND school_mark.section_id='%d' 
        AND school_mark.exam_id='%d' OR school_mark.student_reg_no='%s'  GROUP BY school_mark.student_reg_no''' % (class_id,section_id,exam_id,student_reg_no)
        res = cursor.execute(sql)
        mark_details = cursor.fetchall()
        data ={
            'mark_details':mark_details,
            'student_id': student_id
        }
    else:
        sql= ''' SELECT COUNT(school_mark.student_reg_no),SUM(school_mark.mark),school_mark.student_reg_no,COUNT(school_mark.subject_id) from school_mark where school_mark.class_id='%d' AND school_mark.section_id='%d' 
        AND school_mark.exam_id='%d' AND school_mark.student_reg_no='%s'  GROUP BY school_mark.student_reg_no''' % (class_id,section_id,exam_id,student_reg_no)
        res = cursor.execute(sql)
        mark_details = cursor.fetchall()
        data ={
            'mark_details':mark_details,
            'student_id': student_id
        }
    return JsonResponse(data)

def landing_page(request):
    return render(request,'home.html',{})
from django.db.models.query import QuerySet
def student_class_mark(request,class_id,section_id,exam_id):
    if request.method == "POST":
        ex_id = request.POST.get('ex_id')
        stud_id = request.POST.get('stud_id')
        exam_id = int(ex_id)
        student_id = int(stud_id)
        mark_delete = Mark.objects.filter(exam_id=exam_id,student_reg_no=student_id)
        mark_delete.delete()
    school_id = request.session['schoolname']
    staff_id = request.session['user_id']
    academic_year = request.session['academic_year']
    cursor = connection.cursor()

    stud_ids = ''' SELECT school_mark.student_reg_no, COUNT(school_mark.student_reg_no) from school_mark INNER JOIN
    school_studentdetail ON school_mark.student_reg_no=school_studentdetail.register_number
    WHERE school_mark.class_id='%d' AND school_mark.section_id='%d'
    AND school_mark.exam_id='%d' GROUP BY school_mark.student_reg_no ''' % (class_id,section_id,exam_id)
    exams_id = cursor.execute(stud_ids)
    exams_name = cursor.fetchall()

    a  = Exam.objects.filter(exam_id=exam_id)
    subject_name = AssignSubjectTeacher.objects.filter(class_id=class_id,section_id=section_id,staff_id=staff_id).order_by('subject_id')
    results = Mark.objects.filter(exam_id=exam_id,class_id=class_id,section_id=section_id).select_related('exam').select_related('studentsection')
    cnt = results.values('student_reg_no','exam_id','mark').annotate(count=Count('student_reg_no'),count1=Count('exam_id'))
    page = request.GET.get('page', 1)
    page_count = pagination_cnt(request)
    paginator = Paginator(exams_name,page_count)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    return render(request,'student_class_mark.html',{'exams_name':users,'a':a,'results':cnt,'subject_name':subject_name,'class_id':class_id,
    'section_id':section_id,'exam_id':exam_id})
def mark_already_exist(request):
	cls_id = request.POST.getlist('class_id[]')
	sec_id = request.POST.getlist('section_id[]')
	staff_id = request.POST.getlist('staff_id[]')
	student_id = request.POST.getlist('stud_reg_no[]')
	subject_id = request.POST.getlist('subject_id[]')
	ex_id = request.POST.get('exams')
	exam_id = int(ex_id)
	length = len(subject_id)
	cursor = connection.cursor()
	for i in range(0,length):
		if subject_id[i]!= "":
			post = '''SELECT * from school_mark where school_mark.class_id='%d' AND school_mark.section_id='%d' AND 
			school_mark.subject_id='%d' AND school_mark.exam_id='%d' AND school_mark.staff_id='%d' AND 
			school_mark.student_reg_no='%s' ''' % (int(cls_id[i]),int(sec_id[i]),int(subject_id[i]),exam_id,int(staff_id[i]),student_id[i])
			query = cursor.execute(post)
			row = cursor.fetchall()
			data = {
			'msg':"Mark Already Exist",
			'row':row,
			'length':student_id
			}
			return JsonResponse(data)
def choose_exam(request):
	if request.GET.get('cls_id') and request.GET.get('sec_id'):
		cls_id = request.GET.get('cls_id')
		sec_id = request.GET.get('sec_id')
		class_id = int(cls_id)
		section_id = int(sec_id)
		school_id = request.session['schoolname']
		exams = Exam.objects.filter(class_id_id=class_id,section_id_id=section_id,school_id_id=school_id)
		return render(request,'choose_exam.html',{'exams':exams})
	else:
		return redirect("manage_students_marks")
def exam_mark_chart(request,cls_id,sec_id,pk):
    stud_id = pk
    school_id = request.session['schoolname']
    exams = Exam.objects.filter(class_id_id=cls_id,section_id_id=sec_id,school_id_id=school_id)
    return render(request,'exam_mark_chart.html',{'exams':exams,'pk':pk})

def single_student_exam_mark_chart(request,cls_id,sec_id,pk):
    stud_id = pk
    school_id = request.session['schoolname']
    exams = Exam.objects.filter(class_id_id=cls_id,section_id_id=sec_id,school_id_id=school_id)
    return render(request,'single_student_exam_mark_chart.html',{'exams':exams,'pk':pk})
def exam_chart(request,cls_id,sec_id,exam_id,stud_id):
    exam = Exam.objects.filter(exam_id=exam_id)
    student_name = StudentDetail.objects.filter(register_number=stud_id)
    return render(request,'exam_chart.html',{'exam':exam,'student_name':student_name,'cls_id':cls_id,
    'sec_id':sec_id,'stud_id':stud_id,'exam_id':exam_id})
def particular_exam_chart(request,cls_id,sec_id,exam_id,stud_id):
    exam = Exam.objects.filter(exam_id=exam_id)
    student_name = StudentDetail.objects.filter(register_number=stud_id)
    return render(request,'particular_exam_chart.html',{'exam':exam,'student_name':student_name,'cls_id':cls_id,
    'sec_id':sec_id,'stud_id':stud_id,'exam_id':exam_id})
def sub_ajax(request):
    class_id = request.POST.get('class_id')
    cls_id = int(class_id)
    section_id = request.POST.get('section_id')
    sec_id = int(section_id)
    exam_id = request.POST.get('exam_id')
    ex_id = int(exam_id)
    stud_id = request.POST.get('std_id')
    cursor = connection.cursor()
    sql= ''' SELECT school_subject.subject_name, school_mark.mark from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.class_id='%d' AND school_mark.section_id='%d' 
    AND school_mark.exam_id='%d' AND school_mark.student_reg_no='%s' ''' % (cls_id,sec_id,ex_id,stud_id)
    res = cursor.execute(sql)
    mark_details = cursor.fetchall()

    data = {
    'row':mark_details
    }
    return JsonResponse(data)
def select_exams(request):
	if request.GET.get('cls_id') and request.GET.get('sec_id'):
		cls_id = request.GET.get('cls_id')
		sec_id = request.GET.get('sec_id')
		class_id = int(cls_id)
		section_id = int(sec_id)
		school_id = request.session['schoolname']
		exams = Exam.objects.filter(class_id_id=class_id,section_id_id=section_id,school_id_id=school_id)
		return render(request,'select_exams.html',{'exams':exams})
	else:
		return redirect("manage_students_marks")
def select_exam_excel(request):
    if request.GET.get('cls_id') and request.GET.get('sec_id'):
        cls_id = request.GET.get('cls_id')
        sec_id = request.GET.get('sec_id')
        class_id = int(cls_id)
        section_id = int(sec_id)
        school_id = request.session['schoolname']
        exams = Exam.objects.filter(class_id_id=class_id,section_id_id=section_id,school_id_id=school_id)
        return render(request,'select_exam_excel.html',{'exams':exams})
    else:
        return redirect("manage_students_marks")

def mark_ajax(request):
    class_id = request.POST.get('class_id')
    section_id = request.POST.get('section_id')
    exam_id = request.POST.get('exam_id')
    teacher_id = request.session['user_id']
    cursor = connection.cursor()
    stud_ids = ''' SELECT school_mark.*, school_studentdetail.student_name from school_mark  INNER JOIN school_studentdetail ON 
    school_mark.student_reg_no=school_studentdetail.register_number
    WHERE school_mark.class_id='%d' AND school_mark.section_id='%d'
    AND school_mark.exam_id='%d' AND school_mark.staff_id='%d' order by school_mark.subject_id''' % (int(class_id),int(section_id),int(exam_id),int(teacher_id))
    exams_id = cursor.execute(stud_ids)
    exams_name = cursor.fetchall()


    data = {
    'row':exams_name
    }
    return JsonResponse(data)
def delete_mark(request):
    stud_id = request.POST.get('stud_id')
    exam_id = request.POST.get('ex_id')
    cursor = connection.cursor()
    stud_ids = ''' DELETE from school_mark 
    WHERE school_mark.exam_id='%d' AND school_mark.student_reg_no='%s' ''' % (int(exam_id),stud_id)
    exams_id = cursor.execute(stud_ids)

    data = {
    'row':'deleted'
    }
    return JsonResponse(data)
def view_student_mark(request):
    stud_id = request.session['student_id']
    student = StudentSection.objects.filter(student_id=stud_id)
    return render(request,'view_student_mark.html',{'stud_id':student})
def exam_list(request):
    cls_id = request.POST.get('cls_id')
    sec_id = request.POST.get('sec_id')
    cursor = connection.cursor()
    stud_ids = ''' SELECT * from school_exam 
    WHERE school_exam.class_id_id='%d' AND school_exam.section_id_id='%d'
    order by school_exam.exam_id''' % (int(cls_id),int(sec_id))
    exams_id = cursor.execute(stud_ids)
    exams_name = cursor.fetchall()
    data = {
    'row':exams_name
    }
    return JsonResponse(data)
def view_single_student_particular_exam_mark(request):
    tot = 0
    Register_number = request.session['register_number']
    exam_id =""
    if request.GET.get('id'):
        exam_id = request.GET.get('id')
    cursor = connection.cursor()
    Mark = ''' SELECT school_studentdetail.student_name, school_mark.*,school_mark.subject_id, school_subject.subject_name , school_exam.exams from school_mark
    LEFT JOIN school_subject ON school_mark.subject_id=school_subject.subject_id
    LEFT JOIN school_exam ON school_mark.exam_id=school_exam.exam_id
    LEFT JOIN school_studentdetail ON school_studentdetail.register_number=school_mark.student_reg_no
    WHERE school_mark.exam_id='%d' AND school_mark.student_reg_no='%s'
    ''' % (int(exam_id),Register_number)
    Mark_id = cursor.execute(Mark)
    Marks = cursor.fetchall()
    length = len(Marks)
    for i in Marks:
        tot+=i[3]
    average = tot/length
    if(average>=91 and average<=100):
        grade = "A"
    elif(average>=81 and average<=90):
        grade = "B"
    elif(average>=71 and average<=80):
        grade = "C"
    elif(average>=61 and average<=70):
        grade = "D"
    elif(average>=51 and average<=60):
        grade = "E"
    elif(average>=41 and average<=50):
        grade = "O"
    elif(average>=0 and average<=40):
        grade = "F"

    if grade == "F":
        result = "Fail"
    else:
        result = "Pass"
    percent = format(average, '.1f')
    return render(request,'view_single_student_particular_exam_mark.html',{'Register_number':Register_number,'exam':exam_id,'Marks':Marks,'tot':tot,'length':length,'percentage':percent,'grade':grade,'result':result})
